#
# Example python script to generate a BOM from a KiCad generic netlist
#
# Example: Ungrouped (One component per row) CSV output
#

"""
    @package
    Output: XLSX
    Grouped By: ungrouped, one component per line
    Sorted By: Sheet, then Ref
    Fields: Sheet, Ref, Value, Checkmark, Off-board, Notes

    Command line:
    python "pathToFile/bom_csv_sorted_by_ref.py" "%I" "%O.xlsx"
"""

from __future__ import print_function

# Import the KiCad python helper module
import kicad_netlist_reader
import kicad_utils
import openpyxl
import sys
import re

# A helper function to filter/convert a string read in netlist
#currently: do nothing
def fromNetlistText( aText ):
    return aText

# Generate an instance of a generic netlist, and load the netlist tree from
# the command line option. If the file doesn't exist, execution will stop
net = kicad_netlist_reader.netlist(sys.argv[1])

# Set new workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'BOM'

def writerow( ws, columns ):
    ws.append(columns)

components = net.getInterestingComponents( excludeBOM=True )

components.sort(key=lambda c: (c.getSheetpath(), re.search(r"^[^\d\s]*",c.getRef()).group(0), int(re.search(r"\d*$",c.getRef()).group(0))))

# Output headers
ws.append(['Sheet','Ref', 'Check', 'Value', 'Off-board', 'Notes'] )
ft = openpyxl.styles.Font(bold=True)
for row in ws['A1':'Z1']:
    for cell in row:
        cell.font=ft

# Output all of the component information (One component per row)
for c in components:
    ws.append([c.getSheetpath(), c.getRef(), "", c.getValue(), c.getExcludeFromBoard(), ""])

dimh = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)
for col in range(ws.min_column,ws.max_column+1):
    dimh[openpyxl.utils.get_column_letter(col)] = openpyxl.worksheet.dimensions.ColumnDimension(ws,min=col,max=col,width=20)

ws.column_dimensions = dimh

wb.save(sys.argv[2])

